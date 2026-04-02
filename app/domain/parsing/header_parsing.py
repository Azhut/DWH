"""
Универсальный парсинг заголовков листа (горизонтальные и вертикальные).
Общая логика из BaseSheetParser; используется для 1ФК и 5ФК.
"""
import re
from io import BytesIO
from typing import Any, List, Mapping, Optional

import pandas as pd

from app.domain.parsing.models import ParsedHeaders, TableStructure
from app.domain.parsing.header_fixer import fix_header, finalize_header_fixing
from app.domain.parsing.vertical_hierarchy_config import (
    VerticalHierarchyHeuristicConfig,
    heuristic_config_from_requisites,
    max_vertical_path_segments_from_requisites,
    PATH_SEPARATOR,
)
from app.domain.parsing.workbook_source import ParsingWorkbookSource

_HORIZONTAL_BANNER_SECTION_HINT = re.compile(r"раздел", re.IGNORECASE)
_HORIZONTAL_BANNER_OKEI_HINT = re.compile(r"океи", re.IGNORECASE)


def _strip_leading_segment_if_hint(
    path: str,
    hint: re.Pattern[str],
    *,
    separator: str = PATH_SEPARATOR,
) -> str:
    if path is None:
        return ""
    text = str(path).strip()
    if not text:
        return path if isinstance(path, str) else ""
    parts = [p.strip() for p in text.split(separator)]
    if not parts:
        return text
    out = list(parts)
    if out and hint.search(out[0]):
        out = out[1:]
    if not out:
        return text
    return separator.join(out)


def strip_horizontal_leading_section_banner(path: str, *, separator: str = PATH_SEPARATOR) -> str:
    """
    Снимает первый сегмент пути, если в нём есть «раздел» (шапка раздела в Excel).

    Применяется для **всех** форм после parse_headers.
    """
    return _strip_leading_segment_if_hint(
        path,
        _HORIZONTAL_BANNER_SECTION_HINT,
        separator=separator,
    )


def strip_horizontal_leading_okei_banner(path: str, *, separator: str = PATH_SEPARATOR) -> str:
    """
    Снимает первый сегмент пути, если в нём есть «океи» (блок кодов ОКЕИ).

    Используется **только для 1ФК** (и опционально по реквизиту для других форм).
    Вызывать после ``strip_horizontal_leading_section_banner``.
    """
    return _strip_leading_segment_if_hint(
        path,
        _HORIZONTAL_BANNER_OKEI_HINT,
        separator=separator,
    )


def strip_fk1_horizontal_banner_segments(
    path: str,
    *,
    separator: str = PATH_SEPARATOR,
) -> str:
    """
    Подряд: снять баннер «Раздел», затем «ОКЕИ» — полное правило для 1ФК.

    Для пайплайна эквивалентно: section (всегда) + okei (флаг 1ФК).
    """
    return strip_horizontal_leading_okei_banner(
        strip_horizontal_leading_section_banner(path, separator=separator),
        separator=separator,
    )


def drop_leading_horizontal_path_segments(
    path: str,
    drop_count: int,
    *,
    separator: str = PATH_SEPARATOR,
) -> str:
    """
    Убирает первые ``drop_count`` сегментов иерархического горизонтального заголовка.

    Сегменты разделяются ``PATH_SEPARATOR`` (как после parse_headers / _normalize_headers).
    Не оставляет пустой путь: отбрасывает не больше ``len(segments) - 1`` сегментов.
    """
    if drop_count <= 0 or path is None:
        return path if path is not None else ""
    text = str(path).strip()
    if not text:
        return path if isinstance(path, str) else ""
    parts = text.split(separator)
    if not parts:
        return text
    effective_drop = min(drop_count, max(0, len(parts) - 1))
    if effective_drop <= 0:
        return text
    return separator.join(parts[effective_drop:])


def _get_header_rows(sheet: pd.DataFrame, structure: TableStructure) -> pd.DataFrame:
    """Срезы строк заголовков по структуре."""
    return sheet.iloc[structure.header_start_row : structure.header_end_row + 1].fillna("")


def _fill_empty_cells_in_headers(header_rows: pd.DataFrame, structure: TableStructure) -> None:
    """Заполняет пустые ячейки в заголовках (проброс сверху и слева)."""
    n = len(header_rows)
    for row_idx in range(n - 1, 0, -1):
        for col_idx in range(header_rows.shape[1]):
            if header_rows.iloc[row_idx, col_idx] == "":
                for search_row in range(row_idx - 1, -1, -1):
                    if header_rows.iloc[search_row, col_idx] != "":
                        header_rows.iloc[row_idx, col_idx] = header_rows.iloc[search_row, col_idx]
                        break
    for row_idx in range(n):
        for col_idx in range(1, header_rows.shape[1]):
            if header_rows.iloc[row_idx, col_idx] == "":
                header_rows.iloc[row_idx, col_idx] = header_rows.iloc[row_idx, col_idx - 1]


def _get_horizontal_headers(header_rows: pd.DataFrame) -> List[str]:
    """Формирует горизонтальные заголовки (колонки) из многоуровневых строк."""
    horizontal = []
    n = len(header_rows)
    for col_idx in range(1, header_rows.shape[1]):
        path = []
        current = header_rows.iloc[n - 1, col_idx]
        path.append(current)
        for row_idx in range(n - 2, -1, -1):
            val = header_rows.iloc[row_idx, col_idx]
            if val != current:
                path.insert(0, val)
                current = val
        horizontal.append(PATH_SEPARATOR.join(str(p) for p in path))
    return horizontal


def _get_vertical_header_values_and_row_indices(
    sheet: pd.DataFrame,
    structure: TableStructure,
) -> tuple[List[str], List[int]]:
    """
    Вертикальные заголовки (боковые) из колонки по структуре.

    Возвращает значения и индексы строк DataFrame (нужны для indent-режима).
    """
    col_idx = structure.vertical_header_column
    series = sheet.iloc[structure.data_start_row :, col_idx]
    mask = series.notna()
    values = series[mask].astype(str).tolist()
    df_row_indices = series[mask].index.tolist()
    return values, df_row_indices


def _clean_header_cell(text: object) -> str:
    raw = "" if text is None else str(text)
    return raw.replace("_x000D_", "").strip()


def _cap_stack(stack: List[str], max_path_segments: Optional[int]) -> List[str]:
    if max_path_segments is None or max_path_segments < 1:
        return stack
    if len(stack) <= max_path_segments:
        return stack
    return stack[:1] + stack[-(max_path_segments - 1) :]


def _build_hierarchy_paths_from_depths(
    values: List[str],
    depths: List[int],
    *,
    max_path_segments: Optional[int],
) -> List[str]:
    """
    Строит полный иерархический путь по рассчитанной глубине (0 = корень, далее вложенность).

    max_path_segments: если задано положительное число — обрезка пути до «корень + хвост»;
    None — без ограничения (произвольная рекурсивная глубина).
    """
    stack: List[str] = []
    paths: List[str] = []

    for raw_value, depth in zip(values, depths):
        node = _clean_header_cell(raw_value)
        node = fix_header(node).replace("_x000D_", "").strip()

        depth = max(0, int(depth))
        while len(stack) > depth:
            stack.pop()
        # Скачок глубины (indent / накопленные «в том числе»): восстанавливаем недостающие
        # уровни дублированием последнего известного предка — редкий кейс, но без этого
        # путь обрезался бы до len(stack)+1.
        while len(stack) < depth and stack:
            stack.append(stack[-1])
        if len(stack) < depth and not stack:
            depth = 0
        stack.append(node)
        stack = _cap_stack(stack, max_path_segments)
        paths.append(PATH_SEPARATOR.join(stack))

    return paths


def _subblock_should_exit(stripped: str, sl: str, cfg: VerticalHierarchyHeuristicConfig) -> bool:
    """Конец подблока: новый крупный заголовок (маркеры / длинная строка)."""
    if len(stripped) >= cfg.subblock_exit_min_line_length:
        return True
    return any(marker in sl for marker in cfg.subblock_exit_markers)


def _estimate_depths_heuristic(
    values: List[str],
    cfg: VerticalHierarchyHeuristicConfig,
) -> List[int]:
    """
    Многоуровневые глубины (0, 1, 2, …), в т.ч. рекурсивные «в том числе» подряд.

    Логика:
    - строка, начинающаяся с префикса из primary или compound, наращивает «виток»:
      две такие подряд увеличивают глубину (рекурсия);
    - строка с тире в начале — уровень под текущим витком;
    - ведущие пробелы дают базовый уровень (как прокси indent в .xls);
    - внутри подблока короткие строки без триггеров трактуются как продолжение;
    - длинные / с маркерами («всего») — выход к новому корневому блоку.
    """
    depths: List[int] = []
    phrase_run_depth = 0
    prev_line_was_phrase_opener = False
    in_subblock = False
    last_explicit_depth = 0

    for raw in values:
        s = "" if raw is None else str(raw)
        leading_spaces = len(s) - len(s.lstrip(" "))
        stripped = s.lstrip(" ")
        if not stripped:
            depths.append(0)
            phrase_run_depth = 0
            prev_line_was_phrase_opener = False
            in_subblock = False
            last_explicit_depth = 0
            continue

        sl = stripped.lower()
        base_level = leading_spaces // max(1, cfg.leading_spaces_per_level)

        opens_phrase = any(sl.startswith(p) for p in cfg.primary_child_phrase_prefixes) or any(
            sl.startswith(p) for p in cfg.compound_child_phrase_prefixes
        )
        dash = any(stripped.startswith(d) for d in cfg.dash_prefixes)
        space_only_child = (
            leading_spaces >= cfg.min_leading_spaces_for_child_hint
            and not opens_phrase
            and not dash
        )

        # Выход к новому крупному заголовку (конец подблока)
        exit_block = _subblock_should_exit(stripped, sl, cfg) or (
            in_subblock
            and not opens_phrase
            and not dash
            and not space_only_child
            and leading_spaces == 0
            and len(stripped) >= cfg.subblock_exit_min_line_length
        )
        if in_subblock and exit_block and not opens_phrase and not dash and not space_only_child:
            phrase_run_depth = 0
            prev_line_was_phrase_opener = False
            in_subblock = False
            last_explicit_depth = base_level
            depths.append(base_level)
            continue

        if opens_phrase:
            if prev_line_was_phrase_opener:
                phrase_run_depth += 1
            else:
                phrase_run_depth = max(base_level + 1, 1)
            last_explicit_depth = phrase_run_depth
            depths.append(phrase_run_depth)
            in_subblock = True
            prev_line_was_phrase_opener = True
            continue

        prev_line_was_phrase_opener = False

        if dash:
            d = max(phrase_run_depth, base_level) + 1
            last_explicit_depth = d
            depths.append(d)
            in_subblock = True
            continue

        if space_only_child:
            d = max(phrase_run_depth, base_level) + 1
            last_explicit_depth = d
            depths.append(d)
            in_subblock = True
            continue

        if in_subblock and not _subblock_should_exit(stripped, sl, cfg):
            depths.append(max(last_explicit_depth, phrase_run_depth, base_level))
            continue

        phrase_run_depth = 0
        in_subblock = False
        last_explicit_depth = base_level
        depths.append(base_level)

    if depths:
        baseline = depths[0]
        if baseline > 0:
            depths = [max(0, d - baseline) for d in depths]

    return depths


def _try_get_vertical_indent_levels_from_openpyxl(
    *,
    file_content: bytes,
    sheet_name: str,
    structure: TableStructure,
    values_df_row_indices: List[int],
    vertical_header_column_hint: int = 0,
) -> Optional[List[int]]:
    """
    Пытается прочитать indent-уровни для вертикальной колонки через openpyxl.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        wb = load_workbook(BytesIO(file_content), data_only=True)
    except Exception:
        return None

    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]

    numbering_excel_row = structure.data_start_row
    if numbering_excel_row < 1:
        return None

    first_col_excel: Optional[int] = None
    try:
        for cell in ws[numbering_excel_row]:
            if cell.value is None:
                continue
            if cell.value == 1 or str(cell.value).strip() == "1":
                if first_col_excel is None:
                    first_col_excel = cell.column
                else:
                    first_col_excel = min(first_col_excel, cell.column)
    except Exception:
        return None

    if first_col_excel is None:
        return None

    vertical_header_col_excel = first_col_excel + int(vertical_header_column_hint)

    indent_levels: List[int] = []
    for df_row_idx in values_df_row_indices:
        excel_row = int(df_row_idx) + 1  # df index -> Excel row (1-based)
        try:
            cell = ws.cell(row=excel_row, column=vertical_header_col_excel)
            indent = cell.alignment.indent if cell.alignment else 0
            indent = indent or 0
            indent_levels.append(int(indent))
        except Exception:
            indent_levels.append(0)

    return indent_levels


def _normalize_headers(headers: List[str]) -> List[str]:
    """Удаление переносов и артефактов в заголовках."""
    return [fix_header(h).replace("_x000D_", "").strip() for h in headers]


def parse_headers(
    sheet: pd.DataFrame,
    structure: TableStructure,
    *,
    sheet_name: str = "",
    workbook_source: Optional[ParsingWorkbookSource] = None,
    vertical_hierarchy_mode: str = "auto",
    form_requisites: Mapping[str, Any] | None = None,
) -> ParsedHeaders:
    """
    Парсит горизонтальные и вертикальные заголовки по заданной структуре.

    Вертикальные заголовки строятся как полный путь (root -> ... -> current)
    по иерархии вложенности, разделитель задаётся в vertical_hierarchy_config.PATH_SEPARATOR.

    workbook_source — единый снимок байтов и расширения файла (граница upload → parsing).

    Вызов finalize_header_fixing() — ответственность вызывающего (один раз после всех листов).
    """
    header_rows = _get_header_rows(sheet, structure)
    _fill_empty_cells_in_headers(header_rows, structure)

    horizontal = _get_horizontal_headers(header_rows)

    vertical_values, vertical_df_row_indices = _get_vertical_header_values_and_row_indices(
        sheet,
        structure,
    )

    heuristic_cfg = heuristic_config_from_requisites(form_requisites)
    max_path_segments = max_vertical_path_segments_from_requisites(form_requisites)

    mode = (vertical_hierarchy_mode or "auto").strip().lower()
    if mode not in {"auto", "indent", "heuristics"}:
        mode = "auto"

    ext = ""
    file_content: Optional[bytes] = None
    if workbook_source is not None:
        ext = (workbook_source.extension or "").lower().strip()
        file_content = workbook_source.content

    is_xlsm_like = ext in {".xlsm", "xlsm"}
    is_xlsx_like = ext in {".xlsx", "xlsx"}

    attempt_indent = mode in {"indent", "auto"} and (is_xlsx_like or is_xlsm_like)

    indent_levels: Optional[List[int]] = None
    if attempt_indent and file_content:
        indent_levels = _try_get_vertical_indent_levels_from_openpyxl(
            file_content=file_content,
            sheet_name=sheet_name,
            structure=structure,
            values_df_row_indices=vertical_df_row_indices,
        )

    use_indent = False
    if mode == "indent":
        use_indent = indent_levels is not None
    elif mode == "auto":
        use_indent = indent_levels is not None and any(level > 0 for level in indent_levels)

    if use_indent and indent_levels:
        unique_positive = sorted(set(level for level in indent_levels if level > 0))
        mapping = {level: i + 1 for i, level in enumerate(unique_positive)}
        depths = [0 if level == 0 else mapping.get(level, 1) for level in indent_levels]

        if depths:
            baseline = depths[0]
            if baseline > 0:
                depths = [max(0, d - baseline) for d in depths]

        vertical = _build_hierarchy_paths_from_depths(
            vertical_values,
            depths,
            max_path_segments=max_path_segments,
        )
    else:
        depths = _estimate_depths_heuristic(vertical_values, heuristic_cfg)
        vertical = _build_hierarchy_paths_from_depths(
            vertical_values,
            depths,
            max_path_segments=max_path_segments,
        )

    horizontal = _normalize_headers(horizontal)
    vertical = _normalize_headers(vertical)

    finalize_header_fixing()
    return ParsedHeaders(horizontal=horizontal, vertical=vertical)
