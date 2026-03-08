"""
Helpers to build visual tables from parsed flat_data.

Input contract:
- sheet_data["headers"]["horizontal"]: list[str]
- sheet_data["headers"]["vertical"]: list[str]
- sheet_data["flat_data"]: list[FlatDataRecord|dict]

Backward compatible input (legacy):
- sheet_data["data"] with column_header + values[{row_header, value}]
"""

from __future__ import annotations

import math
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import pandas as pd


def parse_header_levels(header: str, separator: str = " | ") -> List[str]:
    if not header:
        return []
    return [part.strip() for part in str(header).split(separator)]


def _record_get(record: Any, key: str) -> Any:
    if isinstance(record, Mapping):
        return record.get(key)
    return getattr(record, key, None)


def _legacy_to_flat_data(data_columns: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    flat_records: List[Dict[str, Any]] = []
    for column_data in data_columns:
        column_header = column_data.get("column_header")
        for cell in column_data.get("values", []):
            flat_records.append(
                {
                    "row": cell.get("row_header"),
                    "column": column_header,
                    "value": cell.get("value"),
                }
            )
    return flat_records


def _extract_sheet_components(
    sheet_data: Mapping[str, Any],
    *,
    max_rows: int,
    max_cols: int,
) -> tuple[List[str], List[str], List[Any]]:
    headers = sheet_data.get("headers", {})
    horizontal_headers = list(headers.get("horizontal") or [])
    vertical_headers = list(headers.get("vertical") or [])

    flat_records = list(sheet_data.get("flat_data") or [])
    if not flat_records:
        flat_records = _legacy_to_flat_data(sheet_data.get("data") or [])

    if not horizontal_headers:
        horizontal_headers = sorted(
            {
                _record_get(record, "column")
                for record in flat_records
                if _record_get(record, "column") not in (None, "")
            }
        )
    if not vertical_headers:
        vertical_headers = sorted(
            {
                _record_get(record, "row")
                for record in flat_records
                if _record_get(record, "row") not in (None, "")
            }
        )

    return horizontal_headers[:max_cols], vertical_headers[:max_rows], flat_records


def _build_columns_index(horizontal_headers: Sequence[str]) -> tuple[pd.Index | pd.MultiIndex, int]:
    if not horizontal_headers:
        return pd.Index([], name="Headers"), 1

    parsed_levels: List[List[str]] = [parse_header_levels(item) for item in horizontal_headers]
    max_levels = max((len(levels) for levels in parsed_levels), default=1)
    max_levels = max(max_levels, 1)

    if max_levels == 1:
        return pd.Index([levels[0] if levels else "" for levels in parsed_levels], name="Headers"), 1

    level_arrays: List[List[str]] = []
    for level_index in range(max_levels):
        current_level: List[str] = []
        last_value = ""
        for levels in parsed_levels:
            value = levels[level_index] if level_index < len(levels) else ""
            if value == "":
                value = last_value
            else:
                last_value = value
            current_level.append(value)
        level_arrays.append(current_level)

    return (
        pd.MultiIndex.from_arrays(level_arrays, names=[f"Level_{idx + 1}" for idx in range(max_levels)]),
        max_levels,
    )


def build_multiindex_dataframe(
    sheet_data: Dict[str, Any],
    max_rows: int = 100,
    max_cols: int = 50,
) -> Tuple[pd.DataFrame, int, Dict[str, int]]:
    horizontal_headers, vertical_headers, flat_records = _extract_sheet_components(
        sheet_data,
        max_rows=max_rows,
        max_cols=max_cols,
    )

    column_map = {name: index for index, name in enumerate(horizontal_headers)}
    row_map = {name: index for index, name in enumerate(vertical_headers)}

    matrix = [[None] * len(horizontal_headers) for _ in range(len(vertical_headers))]
    diagnostics = {
        "total_flat_records": 0,
        "mapped_records": 0,
        "dropped_unknown_row": 0,
        "dropped_unknown_column": 0,
        "overwritten_cells": 0,
        "empty_values": 0,
    }

    for record in flat_records:
        diagnostics["total_flat_records"] += 1

        row_header = _record_get(record, "row")
        column_header = _record_get(record, "column")
        value = _record_get(record, "value")

        if value in (None, "", "__EMPTY__"):
            diagnostics["empty_values"] += 1

        if row_header not in row_map:
            diagnostics["dropped_unknown_row"] += 1
            continue
        if column_header not in column_map:
            diagnostics["dropped_unknown_column"] += 1
            continue

        row_index = row_map[row_header]
        column_index = column_map[column_header]
        previous = matrix[row_index][column_index]
        if previous is not None and previous != value:
            diagnostics["overwritten_cells"] += 1

        matrix[row_index][column_index] = value
        diagnostics["mapped_records"] += 1

    columns_index, levels_count = _build_columns_index(horizontal_headers)
    dataframe = pd.DataFrame(matrix, index=vertical_headers, columns=columns_index)
    dataframe.index.name = "Rows \\ Columns"

    return dataframe, levels_count, diagnostics


def build_summary_table(
    sheet_data: Dict[str, Any],
    diagnostics: Mapping[str, int] | None = None,
) -> pd.DataFrame:
    horizontal_headers, vertical_headers, flat_records = _extract_sheet_components(
        sheet_data,
        max_rows=10**9,
        max_cols=10**9,
    )

    if diagnostics is None:
        _, _, diagnostics = build_multiindex_dataframe(
            sheet_data,
            max_rows=len(vertical_headers) or 1,
            max_cols=len(horizontal_headers) or 1,
        )

    numeric_values: List[float] = []
    for record in flat_records:
        value = _record_get(record, "value")
        if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
            numeric_values.append(float(value))

    table_cells = len(vertical_headers) * len(horizontal_headers)
    filled_cells = table_cells - diagnostics.get("empty_values", 0)
    fill_percent = (filled_cells / table_cells * 100.0) if table_cells else 0.0

    return pd.DataFrame(
        {
            "Metric": [
                "Flat records (raw)",
                "Header rows",
                "Header columns",
                "Matrix cells",
                "Mapped records",
                "Dropped: unknown row",
                "Dropped: unknown column",
                "Overwritten cells",
                "Empty values",
                "Fill rate",
                "Numeric values",
                "Numeric sum",
                "Numeric mean",
            ],
            "Value": [
                diagnostics.get("total_flat_records", len(flat_records)),
                len(vertical_headers),
                len(horizontal_headers),
                table_cells,
                diagnostics.get("mapped_records", 0),
                diagnostics.get("dropped_unknown_row", 0),
                diagnostics.get("dropped_unknown_column", 0),
                diagnostics.get("overwritten_cells", 0),
                diagnostics.get("empty_values", 0),
                f"{fill_percent:.1f}%",
                len(numeric_values),
                round(sum(numeric_values), 4) if numeric_values else "N/A",
                round(sum(numeric_values) / len(numeric_values), 4) if numeric_values else "N/A",
            ],
        }
    )


def build_flat_data_preview(sheet_data: Dict[str, Any], max_rows: int = 200) -> pd.DataFrame:
    horizontal_headers, vertical_headers, flat_records = _extract_sheet_components(
        sheet_data,
        max_rows=10**9,
        max_cols=10**9,
    )

    row_headers = set(vertical_headers)
    column_headers = set(horizontal_headers)

    rows: List[Dict[str, Any]] = []
    for record in flat_records[:max_rows]:
        row_name = _record_get(record, "row")
        column_name = _record_get(record, "column")
        mapped = row_name in row_headers and column_name in column_headers
        issue = ""
        if not mapped:
            if row_name not in row_headers and column_name not in column_headers:
                issue = "unknown_row+column"
            elif row_name not in row_headers:
                issue = "unknown_row"
            else:
                issue = "unknown_column"

        rows.append(
            {
                "Row": row_name,
                "Column": column_name,
                "Value": _record_get(record, "value"),
                "Mapped": mapped,
                "Issue": issue,
            }
        )

    return pd.DataFrame(rows)


def write_multiindex_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "Data"):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]

    levels_count = df.columns.nlevels if isinstance(df.columns, pd.MultiIndex) else 1
    columns_count = len(df.columns)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    border_thick = Side(style="medium", color="000000")
    border_thin = Side(style="thin", color="BFBFBF")

    header_border = Border(left=border_thick, right=border_thick, top=border_thick, bottom=border_thick)
    data_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    index_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    value_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    if isinstance(df.columns, pd.MultiIndex):
        for level in range(levels_count):
            values = list(df.columns.get_level_values(level))
            for column_idx, value in enumerate(values, start=2):
                worksheet.cell(row=level + 1, column=column_idx, value=str(value))

            start = 0
            for i in range(1, len(values) + 1):
                if i == len(values) or values[i] != values[start]:
                    if i - start > 1:
                        worksheet.merge_cells(
                            start_row=level + 1,
                            start_column=start + 2,
                            end_row=level + 1,
                            end_column=i + 1,
                        )
                    start = i
    else:
        for column_idx, value in enumerate(df.columns, start=2):
            worksheet.cell(row=1, column=column_idx, value=str(value))

    for row in range(1, levels_count + 1):
        for col in range(2, columns_count + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment

    worksheet.cell(row=levels_count + 1, column=1, value=df.index.name or "")
    worksheet.cell(row=levels_count + 1, column=1).font = header_font
    worksheet.cell(row=levels_count + 1, column=1).fill = header_fill
    worksheet.cell(row=levels_count + 1, column=1).alignment = header_alignment
    worksheet.cell(row=levels_count + 1, column=1).border = header_border

    for row_idx, index_value in enumerate(df.index, start=levels_count + 2):
        cell = worksheet.cell(row=row_idx, column=1, value=index_value)
        cell.alignment = index_alignment
        cell.border = header_border

    for row_idx in range(len(df.index)):
        for col_idx in range(columns_count):
            value = df.iloc[row_idx, col_idx]
            if value is None or (isinstance(value, float) and math.isnan(value)):
                value = ""
            cell = worksheet.cell(row=row_idx + levels_count + 2, column=col_idx + 2, value=value)
            cell.alignment = value_alignment
            cell.border = data_border

    worksheet.freeze_panes = worksheet.cell(row=levels_count + 2, column=2)

    end_row = len(df.index) + levels_count + 1
    end_col = columns_count + 1
    if end_row >= levels_count + 1 and end_col >= 1:
        worksheet.auto_filter.ref = f"A{levels_count + 1}:{get_column_letter(end_col)}{end_row}"

    for col_idx in range(1, columns_count + 2):
        width = 18
        if col_idx == 1:
            width = 48
        elif columns_count:
            try:
                sample = [str(df.columns[col_idx - 2])]
            except Exception:
                sample = []
            sample.extend(str(df.iloc[row, col_idx - 2]) for row in range(min(len(df.index), 200)))
            width = max(12, min(55, max(len(item) for item in sample if item is not None) + 2))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row in range(1, levels_count + 1):
        worksheet.row_dimensions[row].height = 36
    for row in range(levels_count + 1, len(df.index) + levels_count + 2):
        worksheet.row_dimensions[row].height = 22

    workbook.save(output_path)

