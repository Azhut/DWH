"""
Утилита для преобразования распарсенных данных в многоуровневую таблицу.
Сохраняет иерархию заголовков (уровни разделены " | ").
"""
import pandas as pd
from typing import List, Dict, Any, Tuple


def parse_header_levels(header: str, separator: str = " | ") -> List[str]:
    """Разбивает заголовок на уровни по разделителю."""
    if not header:
        return []
    parts = [p.strip() for p in header.split(separator)]
    return parts


def build_multiindex_dataframe(
        sheet_data: Dict[str, Any],
        max_rows: int = 100,
        max_cols: int = 50,
) -> Tuple[pd.DataFrame, int]:
    """
    Строит DataFrame с MultiIndex колонками.
    Пустые уровни заполняются последним ненулевым значением (forward-fill).
    """
    horizontal_headers = sheet_data.get("headers", {}).get("horizontal", [])
    vertical_headers = sheet_data.get("headers", {}).get("vertical", [])
    data_columns = sheet_data.get("data", [])

    vertical_headers = vertical_headers[:max_rows]
    horizontal_headers = horizontal_headers[:max_cols]

    # Парсим уровни
    columns_levels: List[List[str]] = []
    max_levels = 0
    for h in horizontal_headers:
        levels = parse_header_levels(h)
        levels = [str(x) if x is not None else "" for x in levels]
        columns_levels.append(levels)
        max_levels = max(max_levels, len(levels))

    if max_levels == 0:
        max_levels = 1
        columns_levels = [[""] for _ in horizontal_headers]

    # Forward-fill внутри каждого столбца
    filled_columns_levels: List[List[str]] = []
    for levels in columns_levels:
        filled = []
        last = ""
        for i in range(max_levels):
            if i < len(levels):
                val = levels[i].strip()
                if val == "":
                    val = last
                else:
                    last = val
            else:
                val = last
            filled.append(val)
        filled_columns_levels.append(filled)

    # Строим level_matrix для MultiIndex
    level_matrix: List[List[str]] = []
    for level_idx in range(max_levels):
        row = []
        last_value = ""
        for col_idx in range(len(filled_columns_levels)):
            val = filled_columns_levels[col_idx][level_idx]
            if not val:
                val = last_value
            else:
                last_value = val
            row.append(val)
        level_matrix.append(row)

    # Создаём MultiIndex
    if max_levels > 1:
        multi_index = pd.MultiIndex.from_arrays(
            level_matrix,
            names=[f"Уровень_{i + 1}" for i in range(max_levels)]
        )
    else:
        multi_index = pd.Index([filled_columns_levels[i][0] for i in range(len(horizontal_headers))], name="Заголовки")

    # Заполняем данными
    data_matrix = [[None] * len(horizontal_headers) for _ in range(len(vertical_headers))]
    for col_data in data_columns:
        col_header = col_data.get("column_header", "")
        if col_header not in horizontal_headers:
            continue
        col_idx = horizontal_headers.index(col_header)
        for cell in col_data.get("values", []):
            row_header = cell.get("row_header", "")
            if row_header not in vertical_headers:
                continue
            row_idx = vertical_headers.index(row_header)
            data_matrix[row_idx][col_idx] = cell.get("value")

    df = pd.DataFrame(data_matrix, index=vertical_headers, columns=multi_index)
    df.index.name = "Строки \\ Колонки"
    return df, max_levels


def build_summary_table(sheet_data: Dict[str, Any]) -> pd.DataFrame:
    """Создаёт сводную таблицу со статистикой по листу."""
    data_columns = sheet_data.get("data", [])
    total_cells = filled_cells = empty_cells = 0
    numeric_sum = 0.0
    numeric_count = 0
    import math

    for col_data in data_columns:
        for cell in col_data.get("values", []):
            total_cells += 1
            value = cell.get("value")
            if value is None or value == "" or value == "__EMPTY__":
                empty_cells += 1
            else:
                filled_cells += 1
                if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
                    numeric_sum += value
                    numeric_count += 1

    return pd.DataFrame({
        "Метрика": ["Всего ячеек", "Заполнено", "Пустых", "% заполненности", "Сумма числовых", "Среднее числовое"],
        "Значение": [
            total_cells, filled_cells, empty_cells,
            f"{(filled_cells / total_cells * 100) if total_cells > 0 else 0:.1f}%",
            round(numeric_sum, 2) if numeric_count > 0 else "N/A",
            round(numeric_sum / numeric_count, 2) if numeric_count > 0 else "N/A",
        ],
    })


def build_flat_data_preview(sheet_data: Dict[str, Any], max_rows: int = 200) -> pd.DataFrame:
    """Создаёт плоский превью данных."""
    flat_rows = []
    for col_data in sheet_data.get("data", []):
        col_header = col_data.get("column_header", "")
        for cell in col_data.get("values", []):
            flat_rows.append({"Row": cell.get("row_header", ""), "Column": col_header, "Value": cell.get("value")})
    return pd.DataFrame(flat_rows).head(max_rows)


def write_multiindex_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "Данные"):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    n_levels = df.columns.nlevels if isinstance(df.columns, pd.MultiIndex) else 1
    ncols = len(df.columns)

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thick = Side(style="medium", color="000000")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    index_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовки
    if isinstance(df.columns, pd.MultiIndex):
        for level in range(n_levels):
            values = df.columns.get_level_values(level)
            for col, val in enumerate(values, start=2):
                ws.cell(row=level + 1, column=col, value=str(val))
    else:
        for col, val in enumerate(df.columns, start=2):
            ws.cell(row=1, column=col, value=str(val))

    # Горизонтальное объединение (только оно стабильно работает)
    if isinstance(df.columns, pd.MultiIndex):
        for level in range(n_levels):
            values = list(df.columns.get_level_values(level))
            start = 0
            for i in range(1, len(values) + 1):
                if i == len(values) or values[i] != values[start]:
                    if i - start > 1:
                        ws.merge_cells(start_row=level + 1, start_column=start + 2, end_row=level + 1, end_column=i + 1)
                    start = i

    # Стили заголовков
    for row in range(1, n_levels + 1):
        for col in range(2, ncols + 2):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

    # Левая колонка
    ws.cell(row=n_levels + 1, column=1, value=df.index.name or "")
    for i, val in enumerate(df.index, start=n_levels + 2):
        ws.cell(row=i, column=1, value=val)
    for row in range(n_levels + 1, len(df.index) + n_levels + 2):
        cell = ws.cell(row=row, column=1)
        cell.alignment = index_alignment
        cell.border = header_border

    # Данные
    import math
    for r in range(len(df.index)):
        for c in range(ncols):
            value = df.iloc[r, c]
            if value is None or (isinstance(value, float) and math.isnan(value)):
                value = ""
            cell = ws.cell(row=r + n_levels + 2, column=c + 2, value=value)
            cell.alignment = data_alignment
            cell.border = data_border

    # Размеры
    for col in range(1, ncols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 55
    for row in range(1, n_levels + 1):
        ws.row_dimensions[row].height = 70
    for row in range(n_levels + 1, len(df.index) + n_levels + 2):
        ws.row_dimensions[row].height = 35

    wb.save(output_path)