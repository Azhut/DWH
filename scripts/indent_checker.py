from openpyxl import load_workbook

# ─── Настройки ────────────────────────────────────────────────
FILE_PATH  = r"C:\Users\Egor\Desktop\Projects\Min_sport\DWH\tests\fixtures\1fk\БИСЕРТЬ 2023.xlsm"  # путь к файлу
# ──────────────────────────────────────────────────────────────


def check_indent(filepath: str) -> None:
    wb = load_workbook(filepath, data_only=True)

    total_found = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_results = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                indent = cell.alignment.indent if cell.alignment else 0
                indent = indent or 0

                if indent == 0:
                    continue

                text = str(cell.value).strip().replace("\n", " ")[:70]
                sheet_results.append((cell.row, cell.column_letter, indent, text))

        if sheet_results:
            print(f"\n{'─' * 70}")
            print(f"  Лист: «{sheet_name}»")
            print(f"{'─' * 70}")
            for row_num, col_letter, indent, text in sheet_results:
                print(f"  строка {row_num:>4}  |  col {col_letter:<3}  |  indent {indent}  |  {text}")
            total_found += len(sheet_results)

    print(f"\n{'═' * 70}")
    if total_found > 0:
        print(f"  ✅ Найдено ячеек с indent: {total_found}")
        print(f"     Иерархию можно читать через cell.alignment.indent")
    else:
        print(f"  ⚠️  Indent не найден ни на одном листе.")
        print(f"     Иерархия, скорее всего, закодирована в тексте")
        print(f"     (пробелы, тире, 'из них' и т.д.) — нужны эвристики.")
    print(f"{'═' * 70}\n")


if __name__ == "__main__":
    check_indent(FILE_PATH)